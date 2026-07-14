"""Versioned parser for the two verified public MinSalud PAI workbooks."""

from __future__ import annotations

import hashlib
import io
import json
import re
import zipfile
from collections.abc import Mapping
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook

from app.ingestion.divipola import normalize_name

PAI_ADAPTER_VERSION = "pai-municipal-v1.1"


@dataclass(frozen=True, slots=True)
class PAIFileContract:
    source_id: str
    url: str
    expected_zip_sha256: str
    expected_workbook_sha256: str
    kind: Literal["history", "current"]
    expected_size: int


PAI_HISTORY_CONTRACT = PAIFileContract(
    source_id="pai-municipal-history",
    url=(
        "https://www.minsalud.gov.co/sites/rid/Lists/BibliotecaDigital/RIDE/VS/PP/PAI/"
        "coberturas-vacunacion-municipal-desde-1998.zip"
    ),
    expected_zip_sha256="5e9d3c8677252f7b88867173ae13600c84901e03c97a6981e8ae74479e8ea2db",
    expected_workbook_sha256="10752884ed45c1e728c1dfb1991acaeaf7e0eae1af51b301328ba82ced3169d4",
    kind="history",
    expected_size=15_368_723,
)

PAI_2026_CONTRACT = PAIFileContract(
    source_id="pai-municipal-2026",
    url=(
        "https://www.minsalud.gov.co/sites/rid/Lists/BibliotecaDigital/RIDE/VS/PP/PAI/"
        "dosis-coberturas-biologicos-municipios-2026.zip"
    ),
    expected_zip_sha256="582ab20337fb52c97798cabd724d6725e7dea56248e153d0bbea9ddfa5e6dc40",
    expected_workbook_sha256="2a224659566c0a9b603cc664f36b456b5be9ad41ef3027eedd78d4e2ce4cbba7",
    kind="current",
    expected_size=2_484_893,
)


@dataclass(frozen=True, slots=True)
class PAIMunicipalRecord:
    municipality_code: str
    year: int
    month: int
    vaccine: str
    source_label: str
    coverage_pct: float
    doses_applied: int | None
    sheet: str
    row_number: int

    def payload(self) -> dict[str, Any]:
        return {
            "municipality_code": self.municipality_code,
            "year": self.year,
            "month": self.month,
            "vaccine": self.vaccine,
            "source_label": self.source_label,
            "coverage_pct": self.coverage_pct,
            "doses_applied": self.doses_applied,
            "sheet": self.sheet,
            "row_number": self.row_number,
        }


@dataclass(frozen=True, slots=True)
class PAIFileRejection:
    row_number: int
    reason_code: str
    reason: str
    payload: dict[str, Any]


@dataclass(slots=True)
class ParsedPAIFile:
    records: list[PAIMunicipalRecord] = field(default_factory=list)
    rejections: list[PAIFileRejection] = field(default_factory=list)
    rows_seen: int = 0
    skipped_rows: int = 0
    workbook_sha256: str | None = None
    schema_descriptor: dict[str, Any] = field(default_factory=dict)
    contract_valid: bool = False


def parse_pai_publication(
    path: str | Path,
    contract: PAIFileContract,
    *,
    start_year: int | None = None,
    end_year: int | None = None,
    months: set[int] | None = None,
    official_municipalities: Mapping[str, str] | None = None,
) -> ParsedPAIFile:
    publication_path = Path(path)
    result = ParsedPAIFile()
    archive_sha256 = _file_sha256(publication_path)
    result.schema_descriptor = {
        "adapter_version": PAI_ADAPTER_VERSION,
        "archive_sha256": archive_sha256,
        "expected_archive_sha256": contract.expected_zip_sha256,
    }
    if archive_sha256 != contract.expected_zip_sha256:
        result.rejections.append(
            PAIFileRejection(
                0,
                "publication_checksum_mismatch",
                "El archivo oficial cambió; se requiere aprobar un nuevo contrato de parser.",
                {
                    "actual_sha256": archive_sha256,
                    "expected_sha256": contract.expected_zip_sha256,
                    "content_bytes": publication_path.stat().st_size,
                },
            )
        )
        return result
    if publication_path.stat().st_size != contract.expected_size:
        result.rejections.append(
            PAIFileRejection(
                0,
                "publication_size_mismatch",
                "El tamaño no coincide con la publicación verificada.",
                {"content_bytes": publication_path.stat().st_size},
            )
        )
        return result

    try:
        with zipfile.ZipFile(publication_path) as archive:
            candidates = [
                item
                for item in archive.infolist()
                if item.filename.lower().endswith(".xlsx")
                and not item.filename.startswith("__MACOSX/")
            ]
            if len(candidates) != 1:
                raise ValueError("El ZIP debe contener exactamente un XLSX principal")
            member = candidates[0]
            if member.file_size > 100 * 1024 * 1024:
                raise ValueError("El XLSX excede el límite de descompresión aprobado")
            workbook_bytes = archive.read(member)
    except (OSError, ValueError, zipfile.BadZipFile) as exc:
        result.rejections.append(
            PAIFileRejection(0, "invalid_zip_contract", str(exc), {"file": "source.zip"})
        )
        return result

    workbook_sha256 = hashlib.sha256(workbook_bytes).hexdigest()
    result.workbook_sha256 = workbook_sha256
    result.schema_descriptor["workbook_sha256"] = workbook_sha256
    result.schema_descriptor["expected_workbook_sha256"] = contract.expected_workbook_sha256
    if workbook_sha256 != contract.expected_workbook_sha256:
        result.rejections.append(
            PAIFileRejection(
                0,
                "workbook_checksum_mismatch",
                "El XLSX interno cambió; se requiere un nuevo adaptador aprobado.",
                {
                    "actual_sha256": workbook_sha256,
                    "expected_sha256": contract.expected_workbook_sha256,
                },
            )
        )
        return result

    workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    sheet_schemas: dict[str, Any] = {}
    selected = _selected_sheets(workbook.sheetnames, contract, start_year, end_year, months)
    for sheet_name, year, month in selected:
        worksheet = workbook[sheet_name]
        header_row = _header_row(sheet_name, contract)
        headers = [cell.value for cell in worksheet[header_row]]
        header_strings = ["" if value is None else str(value) for value in headers]
        header_sha256 = hashlib.sha256(
            json.dumps(header_strings, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
        ).hexdigest()
        measures = _measure_columns(header_strings)
        sheet_schemas[sheet_name] = {
            "header_row": header_row,
            "header_sha256": header_sha256,
            "column_count": len(header_strings),
            "measures": {name: index + 1 for name, index in measures.items()},
        }
        required = {"triple_viral_1y"}
        if year >= 2009:
            required.add("yellow_fever_primary")
        if year >= 2011:
            required.add("influenza_6_11m_second_dose")
        missing = required - measures.keys()
        if missing:
            result.rejections.append(
                PAIFileRejection(
                    header_row,
                    "unknown_header_fingerprint",
                    f"La hoja {sheet_name} no satisface el contrato: {sorted(missing)}",
                    {"sheet": sheet_name, "header_sha256": header_sha256},
                )
            )
            continue
        _parse_sheet(
            worksheet,
            sheet_name=sheet_name,
            year=year,
            month=month,
            header_row=header_row,
            measures=measures,
            result=result,
            contract=contract,
            official_municipalities=official_municipalities,
        )
    workbook.close()
    result.schema_descriptor["sheets"] = sheet_schemas
    result.contract_valid = not any(
        item.reason_code
        in {
            "publication_checksum_mismatch",
            "publication_size_mismatch",
            "invalid_zip_contract",
            "workbook_checksum_mismatch",
            "unknown_header_fingerprint",
            "territory_cardinality_mismatch",
        }
        for item in result.rejections
    )
    return result


def _selected_sheets(
    sheet_names: list[str],
    contract: PAIFileContract,
    start_year: int | None,
    end_year: int | None,
    months: set[int] | None,
) -> list[tuple[str, int, int]]:
    if contract.kind == "history":
        lower = start_year or 1998
        upper = end_year or 2025
        return [
            (name, int(name), 12)
            for name in sheet_names
            if name.isdigit() and lower <= int(name) <= min(upper, 2025)
        ]
    month_names = {"Enero": 1, "Febrero": 2}
    wanted = months or {1, 2}
    return [
        (name, 2026, month_names[name])
        for name in sheet_names
        if name in month_names and month_names[name] in wanted
    ]


def _header_row(sheet_name: str, contract: PAIFileContract) -> int:
    if contract.kind == "current" or sheet_name in {"2008", "2016"}:
        return 6
    return 5


def _identity_indexes(year: int, contract: PAIFileContract) -> tuple[int, int, int]:
    if contract.kind == "current":
        return 0, 2, 3
    if 1998 <= year <= 2003:
        return 2, 4, 5
    if 2004 <= year <= 2008 or 2012 <= year <= 2015 or 2018 <= year <= 2025:
        return 0, 2, 3
    return 1, 3, 4


def _parse_sheet(
    worksheet: Any,
    *,
    sheet_name: str,
    year: int,
    month: int,
    header_row: int,
    measures: dict[str, int],
    result: ParsedPAIFile,
    contract: PAIFileContract,
    official_municipalities: Mapping[str, str] | None,
) -> None:
    dept_index, code_index, name_index = _identity_indexes(year, contract)
    candidates: dict[str, list[tuple[int, tuple[Any, ...]]]] = {}
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        result.rows_seen += 1
        try:
            code, name = _municipality_identity(
                row[dept_index] if dept_index < len(row) else None,
                row[code_index] if code_index < len(row) else None,
                row[name_index] if name_index < len(row) else None,
            )
            if official_municipalities is not None:
                official_name = official_municipalities.get(code)
                if official_name is None:
                    raise ValueError("DIVIPOLA no está en el directorio oficial cargado")
                if year >= 2024 and normalize_name(name) != normalize_name(official_name):
                    # The pinned MinSalud publication contains a finite set of
                    # legacy/short municipality labels (for example, "Cali"
                    # instead of "Santiago de Cali"). The official five-digit
                    # DIVIPOLA code remains the identity key. Preserve every
                    # alias in the immutable schema manifest instead of either
                    # dropping a valid municipality or silently rewriting it.
                    result.schema_descriptor.setdefault("territory_name_aliases", []).append(
                        {
                            "sheet": sheet_name,
                            "municipality_code": code,
                            "source_name": name,
                            "divipola_name": official_name,
                        }
                    )
            candidates.setdefault(code, []).append((row_number, row))
        except ValueError as exc:
            # The official workbooks contain charts, legends and aggregate panels below
            # the municipality block. They are explicitly excluded by the identity
            # contract and reported as skipped, not misrepresented as municipalities.
            _ = exc
            result.skipped_rows += 1

    unique_candidates = {code: rows[0] for code, rows in candidates.items() if len(rows) == 1}
    for code, rows in candidates.items():
        if len(rows) <= 1:
            continue
        for row_number, row in rows:
            result.rejections.append(
                PAIFileRejection(
                    row_number,
                    "duplicate_municipality_identity",
                    "DIVIPOLA aparece más de una vez; no se deduplica silenciosamente.",
                    {
                        "sheet": sheet_name,
                        "municipality_code": code,
                        "municipality_name": _safe_value(row, name_index),
                    },
                )
            )
    if sheet_name in {"2024", "2025", "Enero", "Febrero"} and len(unique_candidates) != 1122:
        result.rejections.append(
            PAIFileRejection(
                header_row,
                "territory_cardinality_mismatch",
                (
                    f"{sheet_name}: se esperaban 1122 municipios y se validaron "
                    f"{len(unique_candidates)}."
                ),
                {"sheet": sheet_name, "validated_municipalities": len(unique_candidates)},
            )
        )
        return

    for code, (row_number, row) in unique_candidates.items():
        for vaccine, column_index in measures.items():
            source_label = str(worksheet.cell(header_row, column_index + 1).value or vaccine)
            raw_coverage = row[column_index] if column_index < len(row) else None
            if raw_coverage is None or str(raw_coverage).strip() == "":
                continue
            try:
                coverage = float(raw_coverage)
                if year <= 2007:
                    coverage *= 100.0
                if not 0 <= coverage <= 150:
                    raise ValueError("coverage_pct fuera del rango administrativo 0..150")
                doses = _optional_int(row[column_index - 1] if column_index else None)
            except (TypeError, ValueError) as exc:
                result.rejections.append(
                    PAIFileRejection(
                        row_number,
                        "invalid_vaccination_measure",
                        str(exc),
                        {
                            "sheet": sheet_name,
                            "municipality_code": code,
                            "measure": vaccine,
                            "raw_coverage": _json_value(raw_coverage),
                        },
                    )
                )
                continue
            result.records.append(
                PAIMunicipalRecord(
                    municipality_code=code,
                    year=year,
                    month=month,
                    vaccine=vaccine,
                    source_label=source_label,
                    coverage_pct=coverage,
                    doses_applied=doses,
                    sheet=sheet_name,
                    row_number=row_number,
                )
            )


def _measure_columns(headers: list[str]) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, header in enumerate(headers):
        normalized = normalize_name(header)
        if not header.strip().startswith("%"):
            continue
        if (
            "REFUERZO" not in normalized
            and "REF." not in normalized
            and (
                "DE TV AL ANO DE EDAD" in normalized
                or "TRIPLE VIRAL AL ANO DE EDAD" in normalized
                or "TRIPLE VIRAL) AL ANO DE EDAD" in normalized
            )
        ):
            result.setdefault("triple_viral_1y", index)
        if "F.A" in normalized and ("1 ANO" in normalized or normalized in {"%F.A.", "% F.A."}):
            result.setdefault("yellow_fever_primary", index)
        if (
            ("FLU" in normalized or "INFLUENZA" in normalized)
            and "6 A 11 MESES" in normalized
            and ("2DA" in normalized or "2DAS" in normalized)
        ):
            result.setdefault("influenza_6_11m_second_dose", index)
    return result


def _municipality_identity(department: Any, code: Any, name: Any) -> tuple[str, str]:
    department_text = _code_text(department)
    code_text = _code_text(code)
    name_text = str(name or "").strip()
    if code_text == "11" and department_text in {"11", "11001"}:
        code_text = "11001"
    elif len(code_text) < 5:
        code_text = code_text.zfill(5)
    if len(code_text) != 5 or not code_text.isdigit():
        raise ValueError("DIVIPOLA/COD debe contener cinco dígitos")
    if code_text == "00000":
        raise ValueError("DIVIPOLA/COD no puede ser 00000")
    department_valid = (
        len(department_text) <= 2 and department_text.zfill(2) == code_text[:2]
    ) or (len(department_text) <= 5 and department_text.zfill(5) == code_text)
    if not department_valid:
        raise ValueError("CODEP/Cons no corresponde al municipio")
    if not name_text or normalize_name(name_text) in {"MUNICIPIO", "TOTAL", "SUBTOTAL"}:
        raise ValueError("el nombre municipal es obligatorio")
    return code_text, name_text


def _code_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\.0$", "", str(value).strip())


def _optional_int(value: Any) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    number = float(value)
    if number < 0:
        raise ValueError("doses_applied no puede ser negativo")
    return int(round(number))


def _safe_value(row: tuple[Any, ...], index: int) -> Any:
    return _json_value(row[index]) if index < len(row) else None


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
