"""Privacy-preserving parser for public INS SIVIGILA 2024 workbooks.

The annual files contain anonymised row-level records and many quasi-identifiers.
This adapter holds those values only while reading the workbook and returns
municipality/week aggregates.  No source row or demographic field is included
in its output, quarantine payloads, snapshots or canonical database writes.
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date
from typing import Any, BinaryIO

from openpyxl import load_workbook

from app.connectors.sivigila_microdata import (
    SIVIGILA2024EventFile,
    SIVIGILAMicrodataMeasure,
)
from app.services.canonical_store import epidemiological_week_start

SIVIGILA_MICRODATA_ADAPTER_VERSION = "sivigila-microdata-2024-v1.0"
MAX_WORKBOOK_ROWS = 2_000_000
MAX_WORKBOOK_COLUMNS = 256
MAX_REJECTION_DETAILS = 250


class SIVIGILAMicrodataContractError(ValueError):
    pass


@dataclass(frozen=True, slots=True)
class SIVIGILASafeRejection:
    row_number: int
    reason_code: str
    reason: str
    safe_payload: dict[str, Any]


@dataclass(frozen=True, slots=True)
class SIVIGILAMunicipalWeek:
    event_code: int
    event_name: str
    disease: str
    municipality_code: str
    epidemiological_year: int
    epidemiological_week: int
    week_start: date
    value: int
    source_rows: int
    measure: str
    canonical_eligible: bool
    territorial_semantics: str

    def snapshot_payload(self) -> dict[str, Any]:
        """Return only the approved aggregate disclosure contract."""

        return {
            "event_code": self.event_code,
            "event_name": self.event_name,
            "disease": self.disease,
            "municipality_code": self.municipality_code,
            "epidemiological_year": self.epidemiological_year,
            "epidemiological_week": self.epidemiological_week,
            "week_start": self.week_start.isoformat(),
            "value": self.value,
            "source_rows": self.source_rows,
            "measure": self.measure,
            "canonical_eligible": self.canonical_eligible,
            "territorial_semantics": self.territorial_semantics,
        }


@dataclass(slots=True)
class ParsedSIVIGILAMicrodata:
    event_code: int
    rows_seen: int = 0
    rows_accepted: int = 0
    rows_rejected: int = 0
    duplicate_rows: int = 0
    records: list[SIVIGILAMunicipalWeek] = field(default_factory=list)
    rejections: list[SIVIGILASafeRejection] = field(default_factory=list)
    schema_descriptor: dict[str, Any] = field(default_factory=dict)


def _header_key(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(character for character in text if not unicodedata.combining(character))
    return re.sub(r"[^A-Z0-9]+", "_", text.upper()).strip("_")


def _integer(value: Any, field_name: str, minimum: int, maximum: int) -> int:
    text = str(value).strip().removesuffix(".0")
    if not text or not re.fullmatch(r"-?\d+", text):
        raise ValueError(f"{field_name} no es un entero")
    number = int(text)
    if number < minimum or number > maximum:
        raise ValueError(f"{field_name} fuera de {minimum}..{maximum}")
    return number


def _code(value: Any, field_name: str, width: int) -> str:
    text = str(value).strip().removesuffix(".0")
    if not text.isdigit() or len(text) > width:
        raise ValueError(f"{field_name} no es un codigo valido")
    return text.zfill(width)


def _required_headers(contract: SIVIGILA2024EventFile) -> set[str]:
    common = {"COD_EVE", "SEMANA", "ANO"}
    if contract.measure == SIVIGILAMicrodataMeasure.IRA_MORBIDITY:
        return common | {
            "NUM_CON",
            "COD_MUN",
            "TOT_IRAG",
            "TOT_IRAUCI",
            "TOT_IRAEXT",
        }
    return common | {"CONSECUTIVE", "COD_DPTO_O", "COD_MUN_O"}


def parse_sivigila_2024_workbook(
    stream: BinaryIO,
    contract: SIVIGILA2024EventFile,
    *,
    max_records: int | None = None,
) -> ParsedSIVIGILAMicrodata:
    """Parse one workbook and immediately collapse it to safe aggregates."""

    stream.seek(0)
    workbook = load_workbook(stream, read_only=True, data_only=True)
    try:
        if not workbook.worksheets:
            raise SIVIGILAMicrodataContractError("El XLSX no contiene hojas")
        worksheet = workbook.worksheets[0]
        if worksheet.max_row > MAX_WORKBOOK_ROWS:
            raise SIVIGILAMicrodataContractError("El XLSX supera el limite de filas")
        if worksheet.max_column > MAX_WORKBOOK_COLUMNS:
            raise SIVIGILAMicrodataContractError("El XLSX supera el limite de columnas")
        iterator = worksheet.iter_rows(values_only=True)
        try:
            source_headers = next(iterator)
        except StopIteration as exc:
            raise SIVIGILAMicrodataContractError("El XLSX esta vacio") from exc
        headers = [_header_key(value) for value in source_headers]
        if not all(headers) or len(headers) != len(set(headers)):
            raise SIVIGILAMicrodataContractError("El XLSX tiene encabezados vacios o duplicados")
        positions = {name: index for index, name in enumerate(headers)}
        required = _required_headers(contract)
        missing = sorted(required - positions.keys())
        if missing:
            raise SIVIGILAMicrodataContractError(
                "Faltan columnas del contrato: " + ", ".join(missing)
            )

        result = ParsedSIVIGILAMicrodata(event_code=contract.event_code)
        aggregates: dict[tuple[str, int], list[int]] = {}
        seen_source_ids: set[str] = set()
        id_column = (
            "NUM_CON"
            if contract.measure == SIVIGILAMicrodataMeasure.IRA_MORBIDITY
            else "CONSECUTIVE"
        )
        territory_semantics = (
            "notifying_municipality"
            if contract.measure == SIVIGILAMicrodataMeasure.IRA_MORBIDITY
            else "municipality_of_occurrence"
        )

        for row_number, row in enumerate(iterator, start=2):
            if max_records is not None and result.rows_seen >= max_records:
                break
            if not any(value is not None and str(value).strip() for value in row):
                continue
            result.rows_seen += 1
            safe_payload: dict[str, Any] = {"event_code": contract.event_code}
            try:
                event_code = _integer(row[positions["COD_EVE"]], "COD_EVE", 1, 999)
                if event_code != contract.event_code:
                    raise ValueError(f"COD_EVE {event_code} no coincide con {contract.event_code}")
                year = _integer(row[positions["ANO"]], "ANO", 2024, 2024)
                week = _integer(row[positions["SEMANA"]], "SEMANA", 1, 53)
                if contract.measure == SIVIGILAMicrodataMeasure.IRA_MORBIDITY:
                    municipality_code = _code(row[positions["COD_MUN"]], "COD_MUN", 5)
                    value = sum(
                        _integer(row[positions[field]], field, 0, 100_000_000)
                        for field in ("TOT_IRAG", "TOT_IRAUCI", "TOT_IRAEXT")
                    )
                else:
                    municipality_code = _code(
                        row[positions["COD_DPTO_O"]], "COD_DPTO_O", 2
                    ) + _code(row[positions["COD_MUN_O"]], "COD_MUN_O", 3)
                    value = 1
                safe_payload.update(
                    {
                        "epidemiological_year": year,
                        "epidemiological_week": week,
                        "municipality_code": municipality_code,
                    }
                )
                source_id = str(row[positions[id_column]] or "").strip()
                if not source_id:
                    raise ValueError(f"{id_column} ausente")
                if source_id in seen_source_ids:
                    result.duplicate_rows += 1
                    raise ValueError(f"{id_column} duplicado")
                seen_source_ids.add(source_id)
                bucket = aggregates.setdefault((municipality_code, week), [0, 0])
                bucket[0] += value
                bucket[1] += 1
                result.rows_accepted += 1
            except (IndexError, TypeError, ValueError) as exc:
                result.rows_rejected += 1
                if len(result.rejections) < MAX_REJECTION_DETAILS:
                    result.rejections.append(
                        SIVIGILASafeRejection(
                            row_number=row_number,
                            reason_code="microdata_row_invalid",
                            reason=str(exc)[:500],
                            safe_payload=safe_payload,
                        )
                    )

        for (municipality_code, week), (value, source_rows) in sorted(aggregates.items()):
            result.records.append(
                SIVIGILAMunicipalWeek(
                    event_code=contract.event_code,
                    event_name=contract.event_name,
                    disease=contract.disease,
                    municipality_code=municipality_code,
                    epidemiological_year=2024,
                    epidemiological_week=week,
                    week_start=epidemiological_week_start(2024, week),
                    value=value,
                    source_rows=source_rows,
                    measure=contract.measure.value,
                    canonical_eligible=contract.canonical_eligible,
                    territorial_semantics=territory_semantics,
                )
            )

        result.schema_descriptor = {
            "adapter_version": SIVIGILA_MICRODATA_ADAPTER_VERSION,
            "source_format": "xlsx",
            "event_code": contract.event_code,
            "source_column_count": len(headers),
            "required_source_columns": sorted(required),
            "persisted_columns": [
                "event_code",
                "event_name",
                "disease",
                "municipality_code",
                "epidemiological_year",
                "epidemiological_week",
                "week_start",
                "value",
                "source_rows",
                "measure",
                "canonical_eligible",
                "territorial_semantics",
            ],
            "privacy_transform": "immediate municipality/week aggregation",
            "raw_rows_persisted": False,
            "quasi_identifiers_persisted": False,
            "canonical_eligible": contract.canonical_eligible,
            "measure": contract.measure.value,
        }
        return result
    finally:
        workbook.close()
